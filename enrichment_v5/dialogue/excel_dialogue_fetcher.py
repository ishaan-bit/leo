"""
Fetch dialogue tuples from Guide to Urban Loneliness.xlsx on HuggingFace Spaces.
Returns 3 randomly selected dialogue tuples from matching domain/secondary row.
"""

import logging
import random
import requests
from typing import Dict, List, Optional, Tuple
from io import BytesIO
from functools import lru_cache
import time

logger = logging.getLogger(__name__)

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("⚠️ openpyxl not installed - Excel dialogue fetching disabled")

# HuggingFace Space URL for Excel file
EXCEL_URL = "https://huggingface.co/spaces/purist-vagabond/enrichment-api-v5/resolve/main/dialogue/Guide%20to%20Urban%20Loneliness.xlsx"

# Domain name mapping (enrichment_v5 → Excel sheet names)
DOMAIN_MAPPING = {
    'self': 'Self',
    'work': 'Work',
    'relationship': 'Relationship',
    'relationships': 'Relationship',
    'health': 'Health',
    'family': 'Family',
    'social': 'Social',
    'study': 'Study',
    'creative': 'Creative',
    'financial': 'Financial',
    'money': 'Financial',
}

# Cache for Excel workbook (refresh every 30 minutes)
_WORKBOOK_CACHE: Optional[openpyxl.Workbook] = None
_CACHE_TIMESTAMP: float = 0
CACHE_TTL = 1800  # 30 minutes


def _fetch_excel_workbook() -> Optional[openpyxl.Workbook]:
    """
    Download and cache the Excel workbook from HuggingFace.
    
    Returns:
        openpyxl.Workbook or None if fetch fails
    """
    global _WORKBOOK_CACHE, _CACHE_TIMESTAMP
    
    if not OPENPYXL_AVAILABLE:
        return None
    
    current_time = time.time()
    
    # Return cached workbook if still valid
    if _WORKBOOK_CACHE and (current_time - _CACHE_TIMESTAMP) < CACHE_TTL:
        logger.debug("Using cached Excel workbook")
        return _WORKBOOK_CACHE
    
    try:
        logger.info(f"📥 Downloading Excel file from {EXCEL_URL}")
        response = requests.get(EXCEL_URL, timeout=15)
        response.raise_for_status()
        
        # Load workbook from bytes
        excel_bytes = BytesIO(response.content)
        workbook = openpyxl.load_workbook(excel_bytes, read_only=True, data_only=True)
        
        # Update cache
        _WORKBOOK_CACHE = workbook
        _CACHE_TIMESTAMP = current_time
        
        logger.info(f"✅ Excel workbook loaded ({len(workbook.sheetnames)} sheets)")
        return workbook
    
    except Exception as e:
        logger.error(f"❌ Failed to fetch Excel file: {e}")
        return None


def _extract_dialogue_tuples_from_row(row: List) -> List[Tuple[str, str, str]]:
    """
    Extract all dialogue tuples from a row.
    
    Expected row format:
    [Domain, Secondary, Dialogue En 1, Dialogue En 2, ..., Dialogue En N, Poem En 1, Poem En 2]
    
    Each Dialogue En cell contains: ["Inner Voice of Reason", "Regulate", "Amuse"]
    
    Args:
        row: Row from Excel sheet
        
    Returns:
        List of tuples (inner_voice, regulate, amuse)
    """
    tuples = []
    
    # Skip first 2 columns (Domain, Secondary)
    # Iterate through remaining columns looking for dialogue tuples
    for cell_idx, cell_value in enumerate(row[2:], start=2):
        if cell_value is None or str(cell_value).strip() == "":
            continue
        
        # Stop at "Poem En" columns
        if "Poem" in str(cell_value):
            break
        
        try:
            # Parse the tuple string
            # Expected format: "['Inner Voice', 'Regulate', 'Amuse']"
            if isinstance(cell_value, str) and cell_value.startswith('['):
                # Use eval safely (only for list literals)
                parsed = eval(cell_value)
                if isinstance(parsed, list) and len(parsed) == 3:
                    tuples.append(tuple(parsed))
            elif isinstance(cell_value, (list, tuple)) and len(cell_value) == 3:
                tuples.append(tuple(cell_value))
        except Exception as e:
            logger.warning(f"Failed to parse dialogue tuple at column {cell_idx}: {e}")
            continue
    
    return tuples


def _extract_poem_from_row(row: List) -> str | None:
    """
    Extract a random poem from Excel row.
    
    Expected row format:
    [Domain, Secondary, Dialogue En 1, ..., Poem En 1, Poem En 2]
    
    Randomly selects between Poem En 1 and Poem En 2.
    
    Args:
        row: Row from Excel sheet
        
    Returns:
        Selected poem text with preserved line breaks, or None if both empty
    """
    poems = []
    
    # Find Poem En 1 and Poem En 2 columns (after dialogue columns)
    for cell_idx, cell_value in enumerate(row):
        # Look for cells after dialogue tuples
        if cell_value is None:
            continue
            
        cell_str = str(cell_value).strip()
        
        # Check if this is a poem column (not a dialogue tuple format)
        # Poems are multi-line strings, not list literals
        if cell_str and not cell_str.startswith('[') and cell_idx >= 2:
            # Skip if this looks like Domain or Secondary column content
            if cell_idx <= 1:
                continue
            
            # This is likely a poem - add to collection
            poems.append(cell_str)
    
    if not poems:
        logger.debug("No poems found in row")
        return None
    
    # Randomly select one poem
    selected_poem = random.choice(poems)
    logger.info(f"Selected poem (length: {len(selected_poem)} chars) from {len(poems)} available")
    
    return selected_poem


def fetch_dialogue_tuples(domain: str, secondary: str) -> Dict:
    """
    Fetch 3 random dialogue tuples from Excel for given domain/secondary.
    
    Args:
        domain: Domain primary (work, self, relationship, etc.)
        secondary: Wheel secondary emotion (Frustrated, Anxious, etc.)
        
    Returns:
        Dict with structure:
        {
            "found": true,
            "domain": "Work",
            "secondary": "Frustrated",
            "tuples": [
                ["Inner Voice 1", "Regulate 1", "Amuse 1"],
                ["Inner Voice 2", "Regulate 2", "Amuse 2"],
                ["Inner Voice 3", "Regulate 3", "Amuse 3"]
            ],
            "source": "excel"
        }
        
        If not found or error:
        {
            "found": false,
            "error": "Description"
        }
    """
    if not OPENPYXL_AVAILABLE:
        return {
            "found": False,
            "error": "openpyxl not installed - cannot read Excel files"
        }
    
    # Normalize inputs
    domain = domain.lower().strip()
    secondary = secondary.strip()  # Keep original casing
    
    # Map domain to sheet name
    sheet_name = DOMAIN_MAPPING.get(domain, 'Self')
    
    # Fetch workbook
    workbook = _fetch_excel_workbook()
    if not workbook:
        return {
            "found": False,
            "error": "Failed to download Excel file from HuggingFace"
        }
    
    # Find sheet
    if sheet_name not in workbook.sheetnames:
        logger.warning(f"Sheet '{sheet_name}' not found in workbook")
        return {
            "found": False,
            "error": f"Sheet '{sheet_name}' not found in Excel file"
        }
    
    sheet = workbook[sheet_name]
    
    # Find row with matching Secondary value
    # Assume: Column A = Domain, Column B = Secondary
    found_row = None
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if len(row) < 2:
            continue
        
        row_secondary = str(row[1]).strip() if row[1] else ""
        
        # Case-insensitive match
        if row_secondary.lower() == secondary.lower():
            found_row = row
            logger.info(f"✅ Found row for {sheet_name}/{secondary} at row {row_idx}")
            break
    
    if not found_row:
        logger.warning(f"❌ No row found for {sheet_name}/{secondary}")
        return {
            "found": False,
            "error": f"No data found for {sheet_name}/{secondary}"
        }
    
    # Extract all dialogue tuples from the row
    all_tuples = _extract_dialogue_tuples_from_row(list(found_row))
    
    if not all_tuples:
        logger.warning(f"No dialogue tuples found in row for {sheet_name}/{secondary}")
        return {
            "found": False,
            "error": f"No dialogue tuples found for {sheet_name}/{secondary}"
        }
    
    # Randomly select 3 tuples (or all if fewer than 3)
    if len(all_tuples) >= 3:
        selected_tuples = random.sample(all_tuples, 3)
    else:
        selected_tuples = all_tuples
        # Pad with fallback if needed
        while len(selected_tuples) < 3:
            selected_tuples.append(("noted.", "breathe.", "pause."))
    
    logger.info(f"🎲 Selected {len(selected_tuples)} dialogue tuples for {sheet_name}/{secondary}")
    
    # Extract poem from same row
    poem = _extract_poem_from_row(list(found_row))
    
    return {
        "found": True,
        "domain": sheet_name,
        "secondary": secondary,
        "tuples": [list(t) for t in selected_tuples],  # Convert to list of lists
        "poem": poem,  # Random selection from Poem En 1 or Poem En 2
        "source": "excel",
        "total_available": len(all_tuples)
    }


# Convenience function for testing
def test_fetch(domain: str = "work", secondary: str = "Frustrated"):
    """Test fetch function with sample data."""
    result = fetch_dialogue_tuples(domain, secondary)
    print(f"\n{'='*60}")
    print(f"Domain: {domain}, Secondary: {secondary}")
    print(f"{'='*60}")
    print(f"Found: {result.get('found')}")
    
    if result.get('found'):
        print(f"Tuples returned: {len(result.get('tuples', []))}")
        for idx, t in enumerate(result.get('tuples', []), 1):
            print(f"\nSet {idx}:")
            print(f"  Inner Voice: {t[0]}")
            print(f"  Regulate: {t[1]}")
            print(f"  Amuse: {t[2]}")
    else:
        print(f"Error: {result.get('error')}")
    print(f"{'='*60}\n")
    
    return result


def build_dialogue_from_excel(
    data: Dict,
    user_id: str = "default_user"
) -> Tuple[List[str], List[str], Dict]:
    """
    Build dialogue (poems + tips) by fetching tuples from Excel file.
    
    This function replaces build_dialogue_from_micro_content() in the pipeline.
    It extracts dialogue tuples from Guide to Urban Loneliness.xlsx instead of batch_all JSON.
    
    Args:
        data: Enrichment result dict with domain.primary, secondary, etc.
        user_id: User identifier (currently unused, kept for API compatibility)
        
    Returns:
        Tuple of (poems, tips, meta)
        - poems: First element of each tuple ("Inner Voice of Reason") - LEGACY
        - tips: Empty list - frontend uses meta['dialogue_tuples'] instead
        - meta: Dict with 'dialogue_tuples' containing full 3-part tuples
        
    Note: Frontend should use meta['dialogue_tuples'] for 3-phase display:
        - tuple[0] = Inner Voice (floating text above city)
        - tuple[1] = Regulate (pig speech bubble)
        - tuple[2] = Amuse (window/building bubble)
    """
    try:
        # Extract domain primary and wheel secondary from enrichment data
        domain = data.get('domain', {})
        raw_domain_primary = domain.get('primary', 'self')
        raw_wheel_secondary = data.get('secondary', 'default')
        
        logger.info(f"[Excel] Raw values: domain={raw_domain_primary}, secondary={raw_wheel_secondary}")
        
        # Fetch from Excel
        excel_result = fetch_dialogue_tuples(raw_domain_primary, raw_wheel_secondary)
        
        if not excel_result.get('found'):
            logger.warning(f"[Excel] ❌ No data found: {excel_result.get('error')}")
            raise ValueError(f"No Excel data found for {raw_domain_primary}/{raw_wheel_secondary}")
        
        # Extract tuples (each tuple is [Inner Voice, Regulate, Amuse])
        tuples = excel_result.get('tuples', [])
        
        if len(tuples) < 3:
            logger.warning(f"[Excel] Only {len(tuples)} tuples available, padding to 3")
            # Pad with fallback tuples
            fallback_tuples = [
                ['Noted.', 'Pause.', 'Breathe.'],
                ['Here.', 'Notice.', 'Allow.'],
                ['Okay.', 'Feel.', 'Continue.']
            ]
            while len(tuples) < 3:
                tuples.append(fallback_tuples[len(tuples)])
        
        # Extract Inner Voice parts as "poems" for backwards compatibility
        # Frontend should use meta['dialogue_tuples'] for the full 3-phase display
        poems = [t[0] for t in tuples[:3]]
        
        # Extract poem (random selection from Poem En 1 or Poem En 2)
        poem = excel_result.get('poem')
        
        logger.info(f"[Excel] ✅ Fetched {len(tuples)} dialogue tuples")
        logger.info(f"[Excel] Tuple 1: {tuples[0] if tuples else 'N/A'}")
        logger.info(f"[Excel] Tuple 2: {tuples[1] if len(tuples) > 1 else 'N/A'}")
        logger.info(f"[Excel] Tuple 3: {tuples[2] if len(tuples) > 2 else 'N/A'}")
        logger.info(f"[Excel] Poem: {'Found' if poem else 'None'} ({len(poem) if poem else 0} chars)")
        
        # Meta information with FULL tuple structure and poem
        meta = {
            'source': 'excel',
            'domain_primary': excel_result.get('domain'),
            'wheel_secondary': excel_result.get('secondary'),
            'total_available': excel_result.get('total_available', len(tuples)),
            'dialogue_tuples': tuples[:3],  # 🔥 NEW: Full 3-part tuples [[Inner, Regulate, Amuse], ...]
            'poem': poem,  # 🔥 NEW: Randomly selected poem from Excel
            'found': True
        }
        
        logger.info(f"Successfully fetched Excel dialogue tuples + poem for {raw_domain_primary}/{raw_wheel_secondary}")
        return poems, [], meta  # tips now empty, frontend uses meta['dialogue_tuples']
    
    except Exception as e:
        logger.error(f"[Excel] ❌ Dialogue fetch failed: {e}")
        logger.exception(e)  # Full stack trace
        
        # Safe fallback with full tuples
        fallback_tuples = [
            ['Noted.', 'Pause.', 'Breathe.'],
            ['Here.', 'Notice.', 'Allow.'],
            ['Okay.', 'Feel.', 'Continue.']
        ]
        
        return (
            ['Noted.', 'Here.', 'Okay.'],  # poems (legacy)
            [],  # tips (empty)
            {
                'source': 'fallback',
                'error': str(e),
                'dialogue_tuples': fallback_tuples,
                'poem': None,  # No poem in fallback
                'found': False
            }
        )


if __name__ == "__main__":
    # Test with sample data
    test_fetch("work", "Frustrated")
    test_fetch("self", "Anxious")
    test_fetch("relationship", "Jealous")
